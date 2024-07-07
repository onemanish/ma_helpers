import streamlit as st
import openpyxl, os, shutil
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from xls2xlsx import XLS2XLSX
print('..............')

# To take the excel output from the 'Compare quotes' in MA, put in the required 
# Formulas in the excel and make it ready for vessel to change qty such that 
# the total amount matches close to required as per SI instructions

def clear_temp_dir():
    if os.path.exists('temp_dir'):
        shutil.rmtree('temp_dir')
    os.makedirs('temp_dir', exist_ok=True)

def quote_maker():
    st.header('MA-QuoteMaker for vessel')
    st.markdown("Upload a Quote file from MA (.xls or .xlsx)")
    uploaded_file = st.file_uploader("Upload a Quote file", type=['xlsx', 'xls'])
    pastels = [ 'FFC3A0', 'A7E6E6', 'FFB6C1', 'A2D4AB', 'F0E68C','D2A3B8', '8FC5E2', 'E6D1A2', 'B8E4C1'] 

    if uploaded_file is not None:
        clear_temp_dir()  # Clear existing files in temp_dir
        st.subheader("Quote details:")
        fileName = os.path.join('temp_dir', uploaded_file.name)
        with open(fileName, "wb") as f:
            f.write(uploaded_file.getbuffer())
        
        #if file is in xls format, convert it to xlsx format - it's not a great conversion though!
        if os.path.splitext(fileName)[1].lower() == '.xls':
            x2x = XLS2XLSX(fileName)
            fileName = os.path.splitext(fileName)[0]+'.xlsx'
            x2x.to_xlsx(filename=fileName)
            
        wbk = openpyxl.load_workbook(fileName)
        if 'rptQuoteDetails' not in wbk.sheetnames: # check whether it is a quote file or not
            st.error('Incorrect quote file format. Please upload quote file from MA')
            return
        sht = wbk['rptQuoteDetails']
        # print('===================')
        
        def get_vendors():
            vendorsTemp = []
            for r in range(13, sht.max_row):
                cellVal = sht.cell(r,1).value
                if cellVal != None:
                    vendorsTemp.append(
                        {'position': cellVal, 'name': sht.cell(r,2).value, 'curr': sht.cell(r,14).value, 'rate': 1.0, 
                        'searchStr': sht.cell(r,2).value + ' (' + sht.cell(r,14).value + ') ' + sht.cell(r,5).value})
                else:
                    break
            return vendorsTemp

        def search_text(rowNum, startCol, searchText):
            resultCol = 0
            for eachCol in range(startCol, sht.max_column+1):
                if sht.cell(rowNum, eachCol).value == searchText:
                    resultCol = eachCol
                    break
            return resultCol    
        
        def get_quote_start(vendorNum):
            quoteStart = 0
            for r in range(12+len(vendorNum),sht.max_row):
                cellVal = sht.cell(r,1).value
                if cellVal == "S.No.": # check for the start of the items section
                    quoteStart = r
                elif quoteStart != 0: # if startRow not found yet then skip to next 
                    break
            return quoteStart + 1

        # get the exchange rate for vendor & currency
        def write_exchange_rates(vendorsTemp):
            curr1col, currUSDcol = 0, 0
            sht.cell(12, search_text(12,1,'Remarks from Vendor')+3, 'Exchange Rate') # write 'exchange rate' beside Vendor table
            for v in range(0, len(vendorsTemp)):
                vendorStartCol = search_text(quoteStartRow-1, 1, vendorsTemp[v]['searchStr']) # look for col start for specific vendor
                # print(f'{vendorStartCol=}')
                curr1col = search_text(quoteStartRow, vendorStartCol, 'Total')
                currUSDcol = search_text(quoteStartRow, vendorStartCol, 'Total(USD)')
                # print(f'{curr1col=}  {currUSDcol=}')
                if currUSDcol > vendorStartCol: # ... if total(USD) col exists > Quote was in other currency
                    curr1 = float(sht.cell(sht.max_row,curr1col).value) # Get val from Total col
                    currUSD = float(sht.cell(sht.max_row,currUSDcol).value) # Get val from Total(USD) col
                    exchRate = curr1/currUSD
                else:
                    exchRate = 1.0
                # print('Write ex rates------', exchRate, vendorsTemp[v]['name'])
                vendorsTemp[v]['rate'] = exchRate # write in revese order as last vendor has left side columns
                sht.cell(13+v,search_text(12,1,'Remarks from Vendor')+3, exchRate) # write exchange rate as above
            return vendorsTemp
        
        def write_formula(vendorsTemp):
            totalCol, totalUSDcol, qtyCol, unitPriceCol, discountAmtCol, vatAmtCol = 0, 0, 0, 0, 0, 0
            # qtyCol = search_text(quoteStartRow, qtyCol+1, 'Qty') # skip the first Qty column
            for v in range(len(vendorsTemp)):
                vendorStartCol = search_text(quoteStartRow-1, 1, vendorsTemp[v]['searchStr']) # look for col start for specific vendor
                totalCol = search_text(quoteStartRow, vendorStartCol, 'Total')
                totalUSDcol = search_text(quoteStartRow, vendorStartCol, 'Total(USD)')
                qtyCol = search_text(quoteStartRow, vendorStartCol, 'Qty')
                unitPriceCol = search_text(quoteStartRow, vendorStartCol, 'Unit Price')
                discountAmtCol = search_text(quoteStartRow, vendorStartCol, 'Discount Amt')
                vatAmtCol = search_text(quoteStartRow, vendorStartCol, 'VAT Amt')
                exRate = vendorsTemp[v]["rate"]
                print(f'{v=}   {vendorsTemp[v]["name"]}: {vendorsTemp[v]["rate"]}')
                # print(f'Column Diff (USD)..... {totalUSDcol-vendorStartCol}')
                for eachRow in range(quoteStartRow+1, sht.max_row):
                    er = str(eachRow) 
                    # formula to be put in each row: Toatl = qty * unit price - discount + VAT
                    totalFormula = f'={get_column_letter(qtyCol)}{er}*{get_column_letter(unitPriceCol)}{er}-{get_column_letter(discountAmtCol)}{er}+{get_column_letter(vatAmtCol)}{er}'
                    sht.cell(eachRow, totalCol, totalFormula)
                    sht[f'{get_column_letter(qtyCol)}{er}'].fill = PatternFill(start_color=pastels[2], end_color=pastels[0], fill_type="solid")
                    if vendorsTemp[v]['curr'] != 'USD':
                        sht.cell(eachRow, totalUSDcol, f'={get_column_letter(totalCol)}{eachRow}/{exRate}')
                
                #Make formula for summing the 'Total' column
                totalColFormula = f'=SUM({get_column_letter(totalCol)}{quoteStartRow+1}:{get_column_letter(totalCol)}{sht.max_row-1})'
                sht.cell(sht.max_row, totalCol, totalColFormula)
                
                # make summing formula for Total(USD) column (it exists only if vendor quote is in other currency)
                if vendorsTemp[v]['curr'] != 'USD':
                    totalUSDColFormula = f'=SUM({get_column_letter(totalUSDcol)}{quoteStartRow+1}:{get_column_letter(totalUSDcol)}{sht.max_row-1})'
                    sht.cell(sht.max_row, totalUSDcol, totalUSDColFormula)         
                    
        def format_sheet(sr):
            for allRows in sht.iter_cols(min_row=sr, max_row=sht.max_row, min_col=3, max_col=sht.max_column):
                for eCell in allRows:
                    eCell.alignment = Alignment(horizontal='center')

        vendors = get_vendors()                     # get number of vendors and details of vendors
        quoteStartRow = get_quote_start(vendors)    # find out which row starts the quotation section
        vendors = write_exchange_rates(vendors)     # find out exchange rates and write to excel next to vendor
        write_formula(vendors)                      # write formula for calculating the item prices and final sum
        format_sheet(quoteStartRow)                 # format the excel sheet nicely
        # save the Excel file in temp_dir, ready for download
        quoteName = sht['A7'].value.replace("Quote Details - ", "").replace(' ','')+'.xlsx'
        fileName = os.path.join('temp_dir', quoteName)
        st.error(quoteName.replace('.xlsx',''))     # not error, just to get it in red color ;-)
        for v in vendors:
            st.info(f"Vendor {v['position']} - {v['name']} - {v['curr']}")
        wbk.save(fileName)
        wbk.close()

        with open(fileName, "rb") as file:
            file_content = file.read()
        st.download_button(
            label="Download Quote File for vessel",
            data=file_content,
            file_name= quoteName,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

def get_quote():
    st.set_page_config(layout='wide', page_title='Quote file')
    quote_maker()

if __name__ == "__main__":
    get_quote()